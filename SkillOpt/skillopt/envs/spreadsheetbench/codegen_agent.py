"""Codegen agent for SpreadsheetBench — no tool-call, pure code generation.

Two modes:
  - **single**: One LLM call → extract ```python``` block → done.
  - **multi**: Up to max_turns LLM calls; after each, execute code and
    feed errors back for correction.

This matches the official SpreadsheetBench evaluation setting (LLM generates
a Python code block, no function-calling / tool-use).
"""
from __future__ import annotations

import json
import os
import random
import signal
import time

import openpyxl


# ── Timeout helper ──────────────────────────────────────────────────────────

class TaskTimeout(Exception):
    """Raised when a task exceeds its time budget."""


def _timeout_handler(signum, frame):
    raise TaskTimeout("Task timed out")

from skillopt.model.azure_openai import (
    get_reasoning_effort,
    get_target_client,
    _needs_responses_api,
    tracker,
)
from skillopt.model import get_codex_exec_config, get_target_backend, is_target_exec_backend
from skillopt.model.codex_harness import prepare_workspace, render_skill_md, run_target_exec
from skillopt.prompts import load_prompt
from skillopt.envs.spreadsheetbench.executor import run_generated_code
from skillopt.envs.spreadsheetbench.evaluator import evaluate


# ── Eval feedback helper (no golden value leakage) ─────────────────────────

def _build_eval_feedback(verify_report: str) -> str:
    """Build Target feedback from a verify report, hiding expected values.

    The verify report contains lines like:
        Sheet1!D2: got=None, expected=0 ✗
        Sheet1!D10: got=None, expected=None ✓

    We strip the ``expected=...`` part so the Target sees only its own
    output and whether each cell is correct or wrong.
    """
    import re
    wrong_lines = []
    n_correct = 0
    for raw_line in verify_report.splitlines():
        raw_line = raw_line.strip()
        if not raw_line:
            continue
        # Match enrichment lines like "  Sheet1!D2: got=None, expected=0 ✗"
        m = re.match(
            r"(\S+!?\w+):\s*got=(.+?),\s*expected=.+?\s*(✓|✗)$",
            raw_line,
        )
        if m:
            cell, got_val, mark = m.groups()
            if mark == "✗":
                wrong_lines.append(f"  {cell}: your output = {got_val} (WRONG)")
            else:
                n_correct += 1
    lines = ["Your code executed successfully but produced incorrect results.",
             "The following cells have wrong values:"]
    lines.extend(wrong_lines)
    if n_correct:
        lines.append(f"  ({n_correct} other cells are correct.)")
    lines.append(
        "\nPlease analyze the spreadsheet data more carefully and fix the code. "
        "Return a complete corrected Python script inside a ```python``` block."
    )
    return "\n".join(lines)


# ── Workbook preview (same as official prompt.py) ────────────────────────────

def _preview_workbook(path: str, max_rows: int = 5, max_cols: int = 20) -> str:
    """Generate a text preview of the first few rows of each sheet."""
    wb = openpyxl.load_workbook(path, data_only=False)
    chunks: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chunks.append(
            f"## Sheet: {sheet_name}  "
            f"(dim={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column})"
        )
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, max_rows),
            max_col=min(ws.max_column, max_cols),
            values_only=False,
        ):
            cells = []
            for cell in row:
                v = cell.value
                if v is None:
                    cells.append(f"{cell.coordinate}=")
                else:
                    s = str(v)
                    if len(s) > 40:
                        s = s[:37] + "..."
                    cells.append(f"{cell.coordinate}={s}")
            chunks.append(" | ".join(cells))
        if ws.max_row > max_rows:
            chunks.append(f"... ({ws.max_row - max_rows} more rows)")
        chunks.append("")
    wb.close()
    return "\n".join(chunks)


# ── Code extraction (same as official prompt.py) ────────────────────────────

def extract_code(text: str) -> str:
    """Extract the first ```python``` fenced code block from LLM output."""
    if "```" not in text:
        return text.strip()
    start = text.find("```")
    nl = text.find("\n", start)
    end = text.find("```", nl + 1)
    if nl == -1 or end == -1:
        return text.strip()
    return text[nl + 1 : end].strip()


# ── Prompt construction (official SpreadsheetBench prompts) ─────────────────


def _build_system(skill_content: str) -> str:
    base = load_prompt("codegen_system", env="spreadsheetbench")
    if skill_content.strip():
        base += f"\n\n## Skill\n{skill_content.strip()}"
    return base


def _build_user(
    instruction: str,
    input_xlsx: str,
    instruction_type: str = "",
    answer_position: str = "",
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
) -> str:
    try:
        preview = _preview_workbook(input_xlsx)
    except Exception as e:  # noqa: BLE001
        preview = f"(failed to preview workbook: {e})"
    extra = ""
    if instruction_type:
        extra += f"\nInstruction type: {instruction_type}"
    if answer_position:
        extra += f"\nExpected answer position: {answer_position}"
    task_suffix = "Return only a ```python``` code block."
    diagnostic = ""
    if diagnostic_mode and diagnostic_instruction.strip():
        task_suffix = (
            "First provide a short diagnostic readout that follows the training "
            "instruction below, then return a single complete ```python``` code block."
        )
        diagnostic = f"\n\n# Training readout\n{diagnostic_instruction.strip()}"
    prefix = ""
    if diagnostic_trace_context.strip():
        prefix = (
            "# Previous Codex Trace Snapshot\n"
            "This is a partial transcript from an earlier attempt. Use it as your current reasoning context.\n\n"
            f"{diagnostic_trace_context.strip()}\n\n"
        )
    return (
        f"{prefix}"
        f"# Instruction\n{instruction}\n{extra}\n\n"
        f"# Input spreadsheet preview\n{preview}\n\n"
        "# Task\n"
        "Write a Python script that reads the workbook from the variable `INPUT_PATH`, "
        "applies the instruction, and writes the modified workbook to `OUTPUT_PATH`. "
        "Preserve all other cells unchanged. "
        "The preview may be truncated — do not hardcode row counts or assume the data ends at the last previewed row; "
        "iterate over all actual rows in the workbook instead. "
        f"{task_suffix}"
        f"{diagnostic}"
    )


# ── LLM call with retry ────────────────────────────────────────────────────

def _llm_call_with_retry(call_fn, *, retries: int = 5, timeout: int | None = 120):
    """Wrap an LLM API call with retry and per-call timeout."""
    last_err = None
    for attempt in range(retries):
        try:
            return call_fn(timeout=timeout)
        except Exception as e:  # noqa: BLE001
            last_err = e
            sleep = min(2 ** attempt + random.random(), 60)
            time.sleep(sleep)
    raise RuntimeError(f"LLM call failed after {retries} retries: {last_err}")


def _get_deployment() -> str:
    from skillopt.model import azure_openai as _llm
    return _llm.TARGET_DEPLOYMENT


def _build_codex_skill(skill_content: str) -> str:
    return render_skill_md(
        skill_content,
        description="Dynamic ReflACT skill for solving the current SpreadsheetBench task.",
        preamble=(
            "Use this skill when solving the current SpreadsheetBench task in this workspace.\n"
            "Write a single self-contained Python solution to `solution.py`.\n"
            "The solution must operate on the provided `INPUT_PATH` and `OUTPUT_PATH` variables.\n"
            "You may inspect `input.xlsx` and run `python run_solution.py` to validate locally,\n"
            "but do not hardcode values from the preview or from one specific workbook."
        ),
    )


def _build_codex_task(
    instruction: str,
    input_xlsx: str,
    instruction_type: str,
    answer_position: str,
    *,
    diagnostic_mode: bool,
    diagnostic_instruction: str,
    diagnostic_trace_context: str,
) -> str:
    prompt = _build_user(
        instruction,
        input_xlsx,
        instruction_type,
        answer_position,
        diagnostic_mode=diagnostic_mode,
        diagnostic_instruction=diagnostic_instruction,
        diagnostic_trace_context=diagnostic_trace_context,
    )
    return (
        f"{prompt}\n\n"
        "## Codex Harness Task\n"
        "- Read `.agents/skills/skillopt-target/SKILL.md` before writing code; do not call a Skill tool.\n"
        "- Read and optionally inspect `input.xlsx` in this workspace.\n"
        "- Write the final Python solution to `solution.py`.\n"
        "- The script should use the provided `INPUT_PATH` and `OUTPUT_PATH` variables.\n"
        "- If you want to validate locally, run `python run_solution.py`.\n"
        "- Do not return a code fence as the primary artifact; the source of truth is `solution.py`.\n"
    )


def _build_codex_driver() -> str:
    return (
        "import pathlib\n"
        "import re\n"
        "import sys\n"
        "import traceback\n\n"
        'INPUT_PATH = "input.xlsx"\n'
        'OUTPUT_PATH = "output.xlsx"\n'
        "code = pathlib.Path('solution.py').read_text(encoding='utf-8')\n"
        "code = re.sub(r'^\\s*(INPUT_PATH|OUTPUT_PATH)\\s*=\\s*.+$', '', code, flags=re.MULTILINE)\n"
        "globals_dict = {'__name__': '__main__', 'INPUT_PATH': INPUT_PATH, 'OUTPUT_PATH': OUTPUT_PATH}\n"
        "try:\n"
        "    exec(compile(code, 'solution.py', 'exec'), globals_dict, globals_dict)\n"
        "except Exception:\n"
        "    traceback.print_exc()\n"
        "    sys.exit(2)\n"
    )


def _prepare_codex_workspace(
    *,
    instruction: str,
    input_xlsx: str,
    output_path: str,
    instruction_type: str,
    answer_position: str,
    skill_content: str,
    diagnostic_mode: bool,
    diagnostic_instruction: str,
    diagnostic_trace_context: str,
    workspace_name: str = "codex_single",
) -> tuple[str, str, str, str]:
    task_out_dir = os.path.dirname(output_path)
    work_dir = os.path.join(task_out_dir, workspace_name)
    skill_md = _build_codex_skill(skill_content)
    task_md = _build_codex_task(
        instruction,
        input_xlsx,
        instruction_type,
        answer_position,
        diagnostic_mode=diagnostic_mode,
        diagnostic_instruction=diagnostic_instruction,
        diagnostic_trace_context=diagnostic_trace_context,
    )
    prompt = (
        "Read `.agents/skills/skillopt-target/SKILL.md` directly; do not call a Skill tool.\n"
        "Read `task.md`, inspect `input.xlsx` if useful, and write the final solution to `solution.py`.\n"
        "You may run `python run_solution.py` to validate the script locally.\n"
        "In your final response, briefly confirm whether `solution.py` was written and summarize the approach."
    )
    prepare_workspace(
        work_dir=work_dir,
        skill_md=skill_md,
        task_text=task_md,
        extra_files={"run_solution.py": _build_codex_driver()},
        copy_files=[(input_xlsx, "input.xlsx")],
    )

    return work_dir, skill_md, task_md, prompt


def _run_exec_backend(
    *,
    work_dir: str,
    prompt: str,
    model: str,
    timeout: int,
) -> tuple[str, str]:
    return run_target_exec(
        work_dir=work_dir,
        prompt=prompt,
        model=model,
        timeout=timeout,
        allow_file_edits=True,
    )


# ── Chat (no tools) ────────────────────────────────────────────────────────

def _chat_call(
    client,
    deployment: str,
    messages: list[dict],
    max_output_tokens: int,
    llm_timeout: int | None = 120,
) -> str:
    """Single LLM call, no tools. Returns raw text."""
    reasoning_effort = get_reasoning_effort()
    if _needs_responses_api(deployment):
        # Responses API
        system = ""
        api_input = []
        for m in messages:
            if m["role"] == "system":
                system = m["content"]
            else:
                api_input.append({"role": m["role"], "content": m["content"]})
        resp = _llm_call_with_retry(lambda timeout: client.responses.create(
            model=deployment,
                instructions=system,
                input=api_input,
                max_output_tokens=max_output_tokens,
                **({"reasoning": {"effort": reasoning_effort}} if reasoning_effort else {}),
                timeout=timeout,
            ), timeout=llm_timeout)
        if hasattr(resp, "usage") and resp.usage:
            tracker.record(
                "rollout",
                getattr(resp.usage, "input_tokens", 0) or 0,
                getattr(resp.usage, "output_tokens", 0) or 0,
            )
        text = getattr(resp, "output_text", None) or ""
        if text:
            return text
        for item in getattr(resp, "output", None) or []:
            for part in getattr(item, "content", []):
                if getattr(part, "type", "") == "output_text":
                    return part.text or ""
        return ""
    else:
        # Chat Completions API — no tools
        kwargs = {
            "model": deployment,
            "messages": messages,
            "max_completion_tokens": max_output_tokens,
        }
        if reasoning_effort is not None:
            kwargs["reasoning_effort"] = reasoning_effort
        resp = _llm_call_with_retry(lambda timeout: client.chat.completions.create(
            **kwargs,
            timeout=timeout,
        ), timeout=llm_timeout)
        if resp.usage:
            tracker.record(
                "rollout",
                resp.usage.prompt_tokens or 0,
                resp.usage.completion_tokens or 0,
            )
        return resp.choices[0].message.content or ""


# ── Public API ──────────────────────────────────────────────────────────────

def run_single(
    instruction: str,
    input_xlsx: str,
    output_path: str,
    instruction_type: str = "",
    answer_position: str = "",
    skill_content: str = "",
    max_output_tokens: int = 16384,
    llm_timeout: int | None = 120,
    task_timeout: int | None = 300,
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
) -> dict:
    """Single-round code generation. One LLM call, no tools.

    Args:
        llm_timeout: Per-LLM-call timeout in seconds (default 120).
        task_timeout: Total task timeout in seconds (default 300).

    Returns ``{"code": str, "raw": str, "n_turns": 1}``.
    """
    no_task_timeout = task_timeout is None or task_timeout <= 0
    if is_target_exec_backend():
        deadline = None if no_task_timeout else time.time() + task_timeout
        deployment = _get_deployment()
        work_dir, skill_md, task_md, prompt = _prepare_codex_workspace(
            instruction=instruction,
            input_xlsx=input_xlsx,
            output_path=output_path,
            instruction_type=instruction_type,
            answer_position=answer_position,
            skill_content=skill_content,
            diagnostic_mode=diagnostic_mode,
            diagnostic_instruction=diagnostic_instruction,
            diagnostic_trace_context=diagnostic_trace_context,
        )
        if deadline is None:
            effective_timeout = 10**9
        else:
            remaining = max(10, int(deadline - time.time()))
            effective_timeout = min(task_timeout, remaining)
        final_message, raw = _run_exec_backend(
            work_dir=work_dir,
            prompt=prompt,
            model=deployment,
            timeout=effective_timeout,
        )
        solution_path = os.path.join(work_dir, "solution.py")
        if os.path.exists(solution_path):
            with open(solution_path, encoding="utf-8") as f:
                code = f.read()
        else:
            code = extract_code(final_message or raw)
        return {
            "code": code,
            "raw": raw or final_message,
            "n_turns": 1,
            "conversation": [{"role": "assistant", "content": final_message or raw}],
            "target_system_prompt": skill_md,
            "target_user_prompt": f"{prompt}\n\n## Task File\n\n{task_md}",
        }

    deadline = None if no_task_timeout else time.time() + task_timeout
    client = get_target_client()
    deployment = _get_deployment()
    system = _build_system(skill_content)
    user = _build_user(
        instruction,
        input_xlsx,
        instruction_type,
        answer_position,
        diagnostic_mode=diagnostic_mode,
        diagnostic_instruction=diagnostic_instruction,
        diagnostic_trace_context=diagnostic_trace_context,
    )

    messages = [
        {"role": "system", "content": system},
        {"role": "user", "content": user},
    ]

    if deadline is None:
        effective_timeout = None
    else:
        remaining = max(10, int(deadline - time.time()))
        effective_timeout = min(llm_timeout or remaining, remaining)
    raw = _chat_call(client, deployment, messages, max_output_tokens, llm_timeout=effective_timeout)
    time.sleep(3)  # Rate-limit cooldown after successful LLM call
    code = extract_code(raw)

    return {
        "code": code,
        "raw": raw,
        "n_turns": 1,
        "conversation": [{"role": "assistant", "content": raw}],
        "target_system_prompt": system,
        "target_user_prompt": user,
    }


def run_multi(
    instruction: str,
    input_xlsx: str,
    output_path: str,
    instruction_type: str = "",
    answer_position: str = "",
    skill_content: str = "",
    max_turns: int = 5,
    max_output_tokens: int = 16384,
    llm_timeout: int | None = 120,
    task_timeout: int | None = 600,
    gold_path: str = "",
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
) -> dict:
    """Multi-round code generation with execution feedback. No tools.

    Each round: LLM generates code → execute → if error, feed back and retry.

    Args:
        llm_timeout: Per-LLM-call timeout in seconds (default 120).
        task_timeout: Total task timeout in seconds (default 600).
        gold_path: Path to golden answer xlsx for eval feedback during
            training.  When non-empty, a successful execution is followed
            by an eval check; if the output is wrong the agent receives
            cell-level feedback (without revealing expected values) and
            gets another turn.  Leave empty for eval/test to avoid
            data leakage.

    Returns ``{"code": str, "raw": str, "n_turns": int, "conversation": [...]}``.
    """
    no_task_timeout = task_timeout is None or task_timeout <= 0
    if is_target_exec_backend():
        deadline = None if no_task_timeout else time.time() + task_timeout
        deployment = _get_deployment()
        work_dir, skill_md, task_md, initial_prompt = _prepare_codex_workspace(
            instruction=instruction,
            input_xlsx=input_xlsx,
            output_path=output_path,
            instruction_type=instruction_type,
            answer_position=answer_position,
            skill_content=skill_content,
            diagnostic_mode=diagnostic_mode,
            diagnostic_instruction=diagnostic_instruction,
            diagnostic_trace_context=diagnostic_trace_context,
            workspace_name="codex_multi",
        )
        prompt = (
            f"{initial_prompt}\n\n"
            "## Multi-Turn Repair Mode\n"
            "- This is turn 1. Write or overwrite `solution.py`.\n"
            "- After each turn, the harness will execute your `solution.py`; if it fails, you will receive feedback and may revise it.\n"
            "- Keep the script general: use `INPUT_PATH` and `OUTPUT_PATH`, and do not hardcode one workbook's values."
        )
        conversation: list[dict] = []
        code = ""
        raw = ""
        final_message = ""
        solution_path = os.path.join(work_dir, "solution.py")

        for turn in range(max_turns):
            if deadline is None:
                effective_timeout = 10**9
            else:
                remaining = deadline - time.time()
                if remaining <= 10:
                    break
                effective_timeout = max(10, int(remaining))
            final_message, raw = _run_exec_backend(
                work_dir=work_dir,
                prompt=prompt,
                model=deployment,
                timeout=effective_timeout,
            )
            conversation.append({"role": "assistant", "content": final_message or raw})

            if os.path.exists(solution_path):
                with open(solution_path, encoding="utf-8") as f:
                    code = f.read()
            else:
                code = extract_code(final_message or raw)
                if code.strip():
                    with open(solution_path, "w", encoding="utf-8") as f:
                        f.write(code)

            if not code.strip():
                feedback = (
                    "No usable `solution.py` or Python code block was produced. "
                    "Write a complete `solution.py` that reads `INPUT_PATH` and saves `OUTPUT_PATH`."
                )
            else:
                ok, err = run_generated_code(
                    code,
                    input_xlsx,
                    output_path,
                    timeout=None if no_task_timeout else 120,
                )
                if ok:
                    if gold_path and answer_position:
                        from skillopt.envs.spreadsheetbench.rollout import _auto_verify_output
                        eval_result = evaluate(
                            output_path, gold_path, instruction_type, answer_position,
                        )
                        if eval_result["ok"]:
                            break
                        verify = _auto_verify_output(output_path, gold_path, answer_position)
                        feedback = _build_eval_feedback(verify)
                    else:
                        break
                else:
                    feedback = (
                        "The current `solution.py` raised an error during harness execution:\n\n"
                        f"```\n{err[:3000]}\n```\n\n"
                        "Revise `solution.py` to fix the error. Keep using `INPUT_PATH` and `OUTPUT_PATH`."
                    )

            feedback_path = os.path.join(work_dir, f"feedback_turn_{turn + 1:02d}.md")
            with open(feedback_path, "w", encoding="utf-8") as f:
                f.write(feedback)
            conversation.append({"role": "user", "content": feedback})
            prompt = (
                f"The previous `solution.py` was evaluated and needs another revision.\n"
                f"Read `{os.path.basename(feedback_path)}` and update `solution.py` accordingly.\n"
                "You may run `python run_solution.py` for a local syntax/runtime check, but the harness will run the final code separately.\n"
                "Do not hardcode workbook-specific answers; preserve unrelated cells."
            )

        return {
            "code": code,
            "raw": raw or final_message,
            "n_turns": len([m for m in conversation if m["role"] == "assistant"]),
            "conversation": conversation,
            "target_system_prompt": skill_md,
            "target_user_prompt": f"{initial_prompt}\n\n## Task File\n\n{task_md}",
        }

    deadline = None if no_task_timeout else time.time() + task_timeout
    client = get_target_client()
    deployment = _get_deployment()
    system = _build_system(skill_content)
    user = _build_user(
        instruction,
        input_xlsx,
        instruction_type,
        answer_position,
        diagnostic_mode=diagnostic_mode,
        diagnostic_instruction=diagnostic_instruction,
        diagnostic_trace_context=diagnostic_trace_context,
    )

    messages: list[dict] = [
        {"role": "system", "content": system},
        {"role": "user", "content": user},
    ]
    conversation: list[dict] = []
    code = ""
    raw = ""

    for turn in range(max_turns):
        if deadline is None:
            effective_timeout = None
        else:
            remaining = deadline - time.time()
            if remaining <= 10:
                # Not enough time for another round
                break
            effective_timeout = min(llm_timeout or int(remaining), int(remaining))
        raw = _chat_call(client, deployment, messages, max_output_tokens, llm_timeout=effective_timeout)
        time.sleep(3)  # Rate-limit cooldown after successful LLM call
        code = extract_code(raw)
        conversation.append({"role": "assistant", "content": raw})
        messages.append({"role": "assistant", "content": raw})

        if not code.strip():
            # No code extracted — ask again
            feedback = (
                "No Python code block was found in your response. "
                "Please return a complete Python script inside a ```python``` block."
            )
            messages.append({"role": "user", "content": feedback})
            conversation.append({"role": "user", "content": feedback})
            continue

        # Execute the code
        ok, err = run_generated_code(
            code,
            input_xlsx,
            output_path,
            timeout=None if no_task_timeout else 120,
        )
        if ok:
            # Execution succeeded — check correctness if gold_path available
            if gold_path and answer_position:
                from skillopt.envs.spreadsheetbench.rollout import _auto_verify_output
                eval_result = evaluate(
                    output_path, gold_path, instruction_type, answer_position,
                )
                if eval_result["ok"]:
                    break  # Genuinely correct — stop

                # Output is wrong — build feedback without leaking golden values
                verify = _auto_verify_output(output_path, gold_path, answer_position)
                feedback = _build_eval_feedback(verify)
                messages.append({"role": "user", "content": feedback})
                conversation.append({"role": "user", "content": feedback})
                continue
            else:
                # No gold path (eval/test) — accept execution success
                break

        # Execution failed — feed error back
        feedback = (
            f"The code raised an error during execution:\n\n"
            f"```\n{err[:3000]}\n```\n\n"
            f"Please fix the code and return a complete corrected Python script "
            f"inside a ```python``` block."
        )
        messages.append({"role": "user", "content": feedback})
        conversation.append({"role": "user", "content": feedback})

    return {
        "code": code,
        "raw": raw,
        "n_turns": turn + 1,
        "conversation": conversation,
        "target_system_prompt": system,
        "target_user_prompt": user,
    }
