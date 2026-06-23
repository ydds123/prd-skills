"""Cell-value evaluator faithful to the official SpreadsheetBench
`evaluation/evaluation.py` (https://github.com/RUCKBReasoning/SpreadsheetBench).

Key rules (copied from the official `transform_value` / `compare_cell_value`):
  * numeric values (int/float and numeric strings) are compared after
    ``round(float(v), 2)`` — a fixed 2-decimal quantization (NOT a tolerance);
  * ``datetime.time`` is stringified and the trailing microseconds stripped;
  * ``datetime.datetime`` is converted to an Excel serial day and rounded
    to an integer day;
  * an empty string ``""`` and ``None`` are considered equal, but otherwise
    ``type(v1) != type(v2)`` fails the comparison.

Format/style comparison is deliberately NOT performed — the official
reference evaluator also skips it (the relevant lines are commented out
in `cell_level_compare`). soft vs hard is defined at the run_bench level
across a task's multiple test cases, not here.
"""
from __future__ import annotations

import datetime
import os
import re

import openpyxl


# ---------- value transform / compare (official port) ----------

def _datetime_to_float(dt: datetime.datetime) -> float:
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def _transform_value(v):
    if isinstance(v, bool):
        # openpyxl can return Python bool; official code doesn't special-case
        # bools, but round(float(True), 2) == 1.0 which breaks 1 vs True. Keep
        # parity with the official transform by promoting bool -> float.
        return round(float(v), 2)
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, datetime.time):
        return str(v)[:-3]
    if isinstance(v, datetime.datetime):
        return round(_datetime_to_float(v), 0)
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    return v


def _compare_cell_value(v1, v2) -> bool:
    v1 = _transform_value(v1)
    v2 = _transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) is not type(v2):
        return False
    return v1 == v2


# ---------- range parsing (official port) ----------

def _col_num2name(n: int) -> str:
    name = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(65 + r) + name
    return name


def _col_name2num(name: str) -> int:
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def _parse_range(range_str: str):
    start_cell, end_cell = range_str.split(":")
    sc = "".join(ch for ch in start_cell if ch.isalpha())
    sr = "".join(ch for ch in start_cell if ch.isdigit())
    ec = "".join(ch for ch in end_cell if ch.isalpha())
    er = "".join(ch for ch in end_cell if ch.isdigit())
    return (_col_name2num(sc), int(sr)), (_col_name2num(ec), int(er))


def _generate_cell_names(range_str: str):
    if ":" not in range_str:
        return [range_str]
    (sc, sr), (ec, er) = _parse_range(range_str)
    cols = [_col_num2name(i) for i in range(sc, ec + 1)]
    return [f"{c}{r}" for c in cols for r in range(sr, er + 1)]


def _cell_level_compare(wb_gt, wb_proc, sheet_name: str, cell_range: str):
    if sheet_name not in wb_proc.sheetnames:
        return False, f"worksheet not found: {sheet_name}"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]
    for cn in _generate_cell_names(cell_range):
        cg = ws_gt[cn]
        cp = ws_proc[cn]
        if not _compare_cell_value(cg.value, cp.value):
            return False, f"value@{sheet_name}!{cn}: gt={cg.value!r} pred={cp.value!r}"
    return True, ""


# ---------- public API ----------

def compare_workbooks(gt_file: str, proc_file: str, answer_position: str) -> tuple[bool, str]:
    """Return (ok, msg). Single test-case comparison, official semantics."""
    if not os.path.exists(proc_file):
        return False, "file not exist"
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:  # noqa: BLE001
        return False, f"load error: {e}"
    try:
        ok_all = True
        msg_first = ""
        for scr in (answer_position or "").split(","):
            scr = scr.strip()
            if not scr:
                continue
            if "!" in scr:
                sheet_name, cell_range = scr.split("!", 1)
                sheet_name = sheet_name.strip().strip("'\"")
            else:
                sheet_name = wb_gt.sheetnames[0]
                cell_range = scr
            cell_range = cell_range.strip().strip("'\"")
            ok, msg = _cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
            if not ok:
                ok_all = False
                if not msg_first:
                    msg_first = msg
        return ok_all, msg_first
    finally:
        wb_gt.close()
        wb_proc.close()


def evaluate(pred_path: str, gold_path: str,
             instruction_type: str, answer_position: str) -> dict:
    """Single test-case evaluate. soft/hard aggregation happens in run_bench."""
    ok, msg = compare_workbooks(gold_path, pred_path, answer_position)
    return {
        "ok": ok,
        "reason": msg,
        "instruction_type": instruction_type,
    }
