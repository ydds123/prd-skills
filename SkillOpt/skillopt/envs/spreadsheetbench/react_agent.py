"""ReAct agent with bash tool for SpreadsheetBench evaluation.

Adapted from the original SpreadsheetBench react agent implementation.

Uses the unified ``skillopt.model`` router so SpreadsheetBench follows the same
backend selection as the rest of the framework.
"""
from __future__ import annotations

import json
import os
import subprocess

from skillopt.model import chat_target_messages
from skillopt.prompts import load_prompt

# ── Tool schemas ─────────────────────────────────────────────────────────────

BASH_TOOL_CHAT = {
    "type": "function",
    "function": {
        "name": "bash",
        "description": (
            "Execute a bash command and receive stdout+stderr (truncated to 4000 chars). "
            "Use Python to read / write Excel files."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "cmd": {"type": "string", "description": "Bash command to execute."}
            },
            "required": ["cmd"],
        },
    },
}

BASH_TOOL_RESPONSES = {
    "type": "function",
    "name": "bash",
    "description": (
        "Execute a bash command and receive stdout+stderr (truncated to 4000 chars). "
        "Use Python to read / write Excel files."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "cmd": {"type": "string", "description": "Bash command to execute."}
        },
        "required": ["cmd"],
    },
}

WRITE_FILE_TOOL_CHAT = {
    "type": "function",
    "function": {
        "name": "write_file",
        "description": (
            "Write content to a file. Use this instead of echo/cat for multi-line "
            "Python scripts to avoid shell escaping issues."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "File path to write (relative to working directory).",
                },
                "content": {
                    "type": "string",
                    "description": "File content to write.",
                },
            },
            "required": ["path", "content"],
        },
    },
}

WRITE_FILE_TOOL_RESPONSES = {
    "type": "function",
    "name": "write_file",
    "description": (
        "Write content to a file. Use this instead of echo/cat for multi-line "
        "Python scripts to avoid shell escaping issues."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "File path to write (relative to working directory).",
            },
            "content": {
                "type": "string",
                "description": "File content to write.",
            },
        },
        "required": ["path", "content"],
    },
}

# ── System prompt ─────────────────────────────────────────────────────────────


def _build_system(skill_content: str) -> str:
    if skill_content.strip():
        skill_section = f"## Skill\n{skill_content.strip()}\n\n"
    else:
        skill_section = ""
    return load_prompt("react_system", env="spreadsheetbench").format(
        critical_rules=load_prompt("critical_rules", env="spreadsheetbench"),
        skill_section=skill_section,
    )


def _build_user(
    instruction: str,
    input_path: str,
    output_path: str,
    instruction_type: str,
    answer_position: str,
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
) -> str:
    parts = []
    if diagnostic_trace_context.strip():
        parts.append(
            "# Previous Codex Trace Snapshot\n"
            "This is a partial transcript from an earlier attempt. Use it as your current reasoning context.\n\n"
            f"{diagnostic_trace_context.strip()}"
        )
    parts.extend([
        f"# Instruction\n{instruction}",
        f"# Input file\n{input_path}",
        f"# Output file\n{output_path}",
    ])
    if instruction_type:
        parts.append(f"# Instruction type\n{instruction_type}")
    if answer_position:
        parts.append(f"# Answer position\n{answer_position}")
    if diagnostic_mode and diagnostic_instruction.strip():
        parts.append(f"# Training readout\n{diagnostic_instruction.strip()}")
    parts.append(
        "Manipulate the input spreadsheet according to the instruction "
        "and save the result to the output file."
    )
    return "\n\n".join(parts)


# ── File write (bypass shell escaping) ────────────────────────────────────────

def _write_file(path: str, content: str, work_dir: str) -> str:
    """Write content to a file, bypassing shell escaping issues."""
    try:
        full_path = os.path.join(work_dir, path) if not os.path.isabs(path) else path
        parent = os.path.dirname(full_path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        with open(full_path, "w") as f:
            f.write(content)
        return f"File written: {full_path} ({len(content)} chars)"
    except Exception as e:  # noqa: BLE001
        return f"[write_file error: {e}]"


# ── Auto-verification ─────────────────────────────────────────────────────────

def _auto_verify(work_dir: str) -> str:
    """Auto-verify output xlsx after solution.py runs."""
    import glob as _glob

    sol_path = os.path.join(work_dir, "solution.py")
    output_path = None
    if os.path.exists(sol_path):
        with open(sol_path) as f:
            for line in f:
                stripped = line.strip()
                if stripped.startswith("OUTPUT_PATH"):
                    try:
                        val = stripped.split("=", 1)[1].strip()
                        output_path = val.strip("'\"").strip()
                    except Exception:  # noqa: BLE001
                        pass
                    break

    if not output_path or not os.path.exists(output_path):
        xlsx_files = [
            f for f in _glob.glob(os.path.join(work_dir, "*.xlsx"))
            if "_pred" in os.path.basename(f)
        ]
        if xlsx_files:
            output_path = xlsx_files[0]

    if not output_path or not os.path.exists(output_path):
        return (
            "\n\n[AUTO-VERIFY] WARNING: Output file not found! "
            "Make sure OUTPUT_PATH is correct and wb.save(OUTPUT_PATH) is called."
        )

    try:
        import openpyxl

        wb_formula = openpyxl.load_workbook(output_path, data_only=False)
        wb_value = openpyxl.load_workbook(output_path, data_only=True)
        lines = [f"\n\n[AUTO-VERIFY] Output file exists: {output_path}"]

        sn = wb_formula.sheetnames[0]
        ws_f = wb_formula[sn]
        ws_v = wb_value[sn]
        lines.append(f"  Sheet '{sn}': {ws_f.dimensions}")

        for row in ws_v.iter_rows(
            min_row=1, max_row=min(5, ws_v.max_row), values_only=True,
        ):
            lines.append(f"    {list(row)}")

        none_cells: list[str] = []
        for row_f, row_v in zip(
            ws_f.iter_rows(min_row=1, max_row=min(30, ws_f.max_row)),
            ws_v.iter_rows(min_row=1, max_row=min(30, ws_v.max_row)),
        ):
            for cf, cv in zip(row_f, row_v):
                formula_val = cf.value
                cached_val = cv.value
                if (
                    isinstance(formula_val, str)
                    and formula_val.startswith("=")
                    and cached_val is None
                ):
                    none_cells.append(cf.coordinate)

        if none_cells:
            lines.append(
                f"  WARNING: {len(none_cells)} cells have formulas but NO cached "
                f"value -- evaluator will see None: {none_cells[:10]}"
            )
            lines.append(
                "  FIX: Compute values in Python and write literal "
                "numbers/strings instead of formulas."
            )
        else:
            lines.append("  All cells have concrete values. Looks good.")

        wb_formula.close()
        wb_value.close()
        return "\n".join(lines)
    except Exception as e:  # noqa: BLE001
        return f"\n\n[AUTO-VERIFY] Could not inspect output: {e}"


# ── Bash execution ────────────────────────────────────────────────────────────

def _run_bash(cmd: str, work_dir: str, timeout: int = 60) -> str:
    try:
        proc = subprocess.run(
            cmd,
            shell=True,
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=work_dir,
        )
        out = (proc.stdout + proc.stderr).strip()
    except subprocess.TimeoutExpired:
        return f"[timeout after {timeout}s]"
    except Exception as e:  # noqa: BLE001
        return f"[error: {e}]"
    if len(out) > 4000:
        out = out[:3800] + f"\n...[truncated, {len(out)} total chars]"
    result = out or "(no output)"

    if "solution.py" in cmd and "python" in cmd.lower():
        result += _auto_verify(work_dir)

    return result


def _assistant_tool_calls(message) -> list[dict]:
    tool_calls = getattr(message, "tool_calls", None) or []
    return [
        tool_call.model_dump() if hasattr(tool_call, "model_dump") else dict(tool_call)
        for tool_call in tool_calls
    ]


def _react_loop(
    system: str,
    user: str,
    work_dir: str,
    max_turns: int,
    max_output_tokens: int,
) -> dict:
    messages: list[dict] = [
        {"role": "system", "content": system},
        {"role": "user", "content": user},
    ]
    conversation: list[dict] = []
    n_turns = 0

    for _ in range(max_turns):
        message, _ = chat_target_messages(
            messages=messages,
            tools=[BASH_TOOL_CHAT, WRITE_FILE_TOOL_CHAT],
            tool_choice="auto",
            max_completion_tokens=max_output_tokens,
            retries=5,
            stage="rollout",
            return_message=True,
        )

        assistant_text = str(getattr(message, "content", "") or "")
        tool_calls = _assistant_tool_calls(message)
        assistant_payload: dict = {"role": "assistant", "content": assistant_text}
        if tool_calls:
            assistant_payload["tool_calls"] = tool_calls
        messages.append(assistant_payload)

        if not tool_calls:
            conversation.append({"type": "message", "content": assistant_text})
            break

        for tool_call in tool_calls:
            n_turns += 1
            function = tool_call.get("function", {}) or {}
            try:
                args = json.loads(str(function.get("arguments", "{}") or "{}"))
            except json.JSONDecodeError:
                args = {}

            if function.get("name") == "write_file":
                obs = _write_file(
                    args.get("path", ""),
                    args.get("content", ""),
                    work_dir,
                )
                conversation.append({
                    "type": "tool_call",
                    "cmd": f"[write_file] {args.get('path', '')}",
                    "obs": obs,
                })
            else:
                cmd = args.get("cmd", "")
                obs = _run_bash(cmd, work_dir)
                conversation.append({"type": "tool_call", "cmd": cmd, "obs": obs})

            messages.append(
                {
                    "role": "tool",
                    "tool_call_id": tool_call.get("id", ""),
                    "content": obs,
                }
            )

    return {"conversation": conversation, "n_turns": n_turns}


# ── Public API ────────────────────────────────────────────────────────────────

def run_react(
    instruction: str,
    input_path: str,
    output_path: str,
    work_dir: str,
    instruction_type: str = "",
    answer_position: str = "",
    skill_content: str = "",
    max_turns: int = 30,
    max_output_tokens: int = 16384,
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
) -> dict:
    """Run the ReAct agent for one task.

    Returns:
        {
          "conversation": [...],  # list of {type, cmd/content, obs?}
          "n_turns": int,         # number of bash tool calls made
        }
    """
    system = _build_system(skill_content)
    user = _build_user(
        instruction,
        input_path,
        output_path,
        instruction_type,
        answer_position,
        diagnostic_mode=diagnostic_mode,
        diagnostic_instruction=diagnostic_instruction,
        diagnostic_trace_context=diagnostic_trace_context,
    )
    result = _react_loop(system, user, work_dir, max_turns, max_output_tokens)
    result["target_system_prompt"] = system
    result["target_user_prompt"] = user
    return result
