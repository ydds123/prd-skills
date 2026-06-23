"""SpreadsheetBench rollout — codegen & ReAct batch execution.

Provides:
  - process_one_codegen(): single/multi-round code generation (no tool-call)
  - run_spreadsheet_batch_codegen(): batch wrapper for codegen
  - process_one(): ReAct agent with tool-call (legacy)
  - run_spreadsheet_batch(): batch wrapper for ReAct (legacy)
  - load_items(): load benchmark .json/.jsonl files
"""
from __future__ import annotations

import glob as _glob
import json
import os
import shutil
import tempfile
import time
import traceback
from concurrent.futures import (
    FIRST_COMPLETED,
    ThreadPoolExecutor,
    wait,
    TimeoutError as FuturesTimeoutError,
)

import openpyxl

from skillopt.envs.spreadsheetbench.react_agent import run_react
from skillopt.envs.spreadsheetbench.evaluator import (
    evaluate, _generate_cell_names, _compare_cell_value,
)
from skillopt.envs.spreadsheetbench.executor import run_generated_code


# ── Data loading ─────────────────────────────────────────────────────────────


def load_items(path: str) -> list[dict]:
    """Load a benchmark file. Supports both .jsonl and .json (list of dicts)."""
    if path.endswith(".json"):
        with open(path) as f:
            data = json.load(f)
        if isinstance(data, dict):
            data = data.get("data") or list(data.values())
        return list(data)
    items = []
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line:
                items.append(json.loads(line))
    return items


# ── Test case discovery ──────────────────────────────────────────────────────


def _find_test_cases(task_dir: str) -> list[tuple[str, str, str]]:
    """Return [(case_no, input_path, answer_path), ...] sorted by case_no.

    Supports naming conventions used by SpreadsheetBench releases:
      * ``{no}_{id}_input.xlsx``  + ``{no}_{id}_answer.xlsx``  (original)
      * ``{no}_{id}_init.xlsx``   + ``{no}_{id}_golden.xlsx``  (verified_400)
      * ``initial.xlsx``          + ``golden.xlsx``             (verified_400, no prefix)
    """
    cases: list[tuple[str, str, str]] = []
    inputs = sorted(_glob.glob(os.path.join(task_dir, "*_input.xlsx")))
    for ip in inputs:
        no = os.path.basename(ip).split("_", 1)[0]
        ap = ip.replace("_input.xlsx", "_answer.xlsx")
        if os.path.exists(ap):
            cases.append((no, ip, ap))
    inits = sorted(_glob.glob(os.path.join(task_dir, "*_init.xlsx")))
    for ip in inits:
        no = os.path.basename(ip).split("_", 1)[0]
        ap = ip.replace("_init.xlsx", "_golden.xlsx")
        if os.path.exists(ap):
            cases.append((no, ip, ap))

    # Fallback: bare initial.xlsx + golden.xlsx (no numbered prefix)
    if not cases:
        bare_init = os.path.join(task_dir, "initial.xlsx")
        bare_gold = os.path.join(task_dir, "golden.xlsx")
        if os.path.exists(bare_init) and os.path.exists(bare_gold):
            cases.append(("1", bare_init, bare_gold))

    return cases


# ── Auto-verify helper ──────────────────────────────────────────────────────

# The official SpreadsheetBench evaluator never serialises cells to text — it
# compares in memory and returns only a pass/fail bool. The per-cell report
# below is a repo-local training aid (fed back to the model on retry and saved
# into the trajectory for reflection). On most tasks the answer range is a
# handful of cells, so the full report is tiny. But a few tasks have answer
# ranges spanning tens of thousands of cells (e.g. 80-42 =
# 'Consolidate_ALL'!A2:L8000 ≈ 96k cells); dumping every cell explodes the
# report to several MB, floods the model's context and bloats conversation
# files. We therefore apply the same head+tail character truncation the rest of
# the codebase uses for oversized trajectory text (cf. reflect.py / slow_update.py
# `text[:half] + "...[truncated]...\n" + text[-half:]`): keep the first and last
# `_MAX_REPORT_CHARS // 2` chars so both the leading and trailing wrong cells
# stay visible. Small reports are unchanged.
_MAX_REPORT_CHARS = 12000      # head+tail char budget (~6000 head + 6000 tail)


def _auto_verify_output(
    pred_path: str,
    gold_path: str,
    answer_position: str,
) -> str:
    """Reopen the predicted xlsx and compare cells at answer_position with gold.

    Returns a human-readable verification report that can be appended to the
    trajectory so the error analyst can see exactly what went wrong (e.g.
    ``cell A1: got=None, expected=420``). Oversized reports are head+tail
    truncated to `_MAX_REPORT_CHARS` chars, matching the rest of the codebase.
    """
    if not os.path.exists(pred_path):
        return "Verification: output file does not exist."
    try:
        wb_pred = openpyxl.load_workbook(pred_path, data_only=True)
        wb_gold = openpyxl.load_workbook(gold_path, data_only=True)
    except Exception as e:
        return f"Verification: could not open workbooks: {e}"

    lines = ["## Output Verification"]
    try:
        for scr in (answer_position or "").split(","):
            scr = scr.strip()
            if not scr:
                continue
            if "!" in scr:
                sheet_name, cell_range = scr.split("!", 1)
                sheet_name = sheet_name.strip().strip("'\"")
            else:
                sheet_name = wb_gold.sheetnames[0]
                cell_range = scr
            cell_range = cell_range.strip().strip("'\"")

            cell_names = _generate_cell_names(cell_range)
            ws_pred = wb_pred[sheet_name] if sheet_name in wb_pred.sheetnames else None
            ws_gold = wb_gold[sheet_name] if sheet_name in wb_gold.sheetnames else None

            if ws_pred is None:
                lines.append(f"  Sheet '{sheet_name}' NOT FOUND in output.")
                continue

            n_empty_correct = 0   # empty-on-both correct cells collapsed to a count
            for cn in cell_names:
                gv = ws_gold[cn].value if ws_gold else "N/A"
                pv = ws_pred[cn].value
                # Use the official cell comparator so this report's ✓/✗ agrees
                # with the real scorer (evaluate). repr() equality would wrongly
                # flag e.g. 5 vs 5.0 or None vs "" as mismatches and mislead the
                # model into "fixing" cells that already pass scoring.
                ok_cell = ws_gold is not None and _compare_cell_value(gv, pv)
                # Collapse only cells that are correct AND empty on both sides
                # (got=None, expected=None ✓): pure noise. Every other cell —
                # including non-empty correct cells — is listed in full; the
                # final head+tail char cap keeps the report bounded.
                if ok_cell and gv in (None, "") and pv in (None, ""):
                    n_empty_correct += 1
                    continue
                match = "✓" if ok_cell else "✗"
                lines.append(f"  {sheet_name}!{cn}: got={pv!r}, expected={gv!r} {match}")
            if n_empty_correct:
                lines.append(
                    f"  (+{n_empty_correct} empty cells correct, omitted)"
                )

        # Also check if any cells in the output contain formula strings
        formula_cells = []
        for sn in wb_pred.sheetnames:
            ws = wb_pred[sn]
            for row in ws.iter_rows(max_row=min(ws.max_row, 200), values_only=False):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells.append(f"{sn}!{cell.coordinate}={cell.value}")
                        if len(formula_cells) >= 10:
                            break
                if len(formula_cells) >= 10:
                    break
            if len(formula_cells) >= 10:
                break
        if formula_cells:
            lines.append(f"\n  WARNING: {len(formula_cells)} cells contain Excel formulas (openpyxl cannot evaluate them):")
            for fc in formula_cells[:5]:
                lines.append(f"    {fc}")
            if len(formula_cells) > 5:
                lines.append(f"    ... and {len(formula_cells) - 5} more")
    finally:
        wb_pred.close()
        wb_gold.close()

    report = "\n".join(lines)
    # Head+tail truncation, matching reflect.py / slow_update.py: keep the first
    # and last half so both leading and trailing wrong cells remain visible.
    if len(report) > _MAX_REPORT_CHARS:
        half = _MAX_REPORT_CHARS // 2
        report = (
            report[:half]
            + f"\n  ...[verification report truncated, {len(report)} chars total]...\n"
            + report[-half:]
        )
    return report


# ── Per-task worker ──────────────────────────────────────────────────────────


def process_one(
    item: dict,
    data_root: str,
    out_root: str,
    skill_content: str,
    max_turns: int,
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
    max_completion_tokens: int = 16384,
) -> dict:
    """Run the ReAct agent on a single SpreadsheetBench task.

    Returns a result dict compatible with ``compute_score()``.
    """
    task_id = str(item["id"])
    instruction = item["instruction"]
    instruction_type = item.get("instruction_type", "")
    answer_position = item.get("answer_position", "")
    answer_sheet = item.get("answer_sheet", "")
    if answer_position and answer_sheet and "!" not in answer_position:
        answer_position_eval = f"{answer_sheet}!{answer_position}"
    else:
        answer_position_eval = answer_position

    # Determine task_type from instruction_type
    itype_lower = (instruction_type or "").lower()
    if "cell" in itype_lower:
        task_type = "cell_level"
    elif "sheet" in itype_lower:
        task_type = "sheet_level"
    else:
        task_type = "other"

    sp = item.get("spreadsheet_path", f"spreadsheet/{task_id}")
    task_dir = sp if os.path.isabs(sp) else os.path.join(data_root, sp)

    result = {
        "id": task_id,
        "ok": False,
        "instruction_type": instruction_type,
        "task_type": task_type,
        "task_description": instruction,
        "phase": "setup",
        "fail_reason": "",
        "agent_ok": False,
        "exec_ok": False,
        "n_cases": 0,
        "n_exec_pass": 0,
        "n_pass": 0,
        "soft": 0.0,
        "hard": 0,
        "n_turns": 0,
        "cases": [],
        "error": "",
    }

    try:
        cases = _find_test_cases(task_dir)
        result["n_cases"] = len(cases)
        if not cases:
            result["fail_reason"] = "no-test-cases"
            return result

        task_out_dir = os.path.join(out_root, "predictions", task_id)
        os.makedirs(task_out_dir, exist_ok=True)

        no1, ip1, _ = cases[0]
        pred_path_1 = os.path.join(task_out_dir, f"{no1}_pred.xlsx")
        target_prompt_parts = [
            f"# Instruction\n{instruction}",
            f"# Input file\n{ip1}",
            f"# Output file\n{pred_path_1}",
        ]
        if instruction_type:
            target_prompt_parts.append(f"# Instruction type\n{instruction_type}")
        if answer_position_eval:
            target_prompt_parts.append(f"# Answer position\n{answer_position_eval}")
        if diagnostic_trace_context.strip():
            target_prompt_parts.insert(
                0,
                "# Previous Codex Trace Snapshot\n"
                "This is a partial transcript from an earlier attempt. Use it as your current reasoning context.\n\n"
                f"{diagnostic_trace_context.strip()}",
            )
        if diagnostic_mode and diagnostic_instruction.strip():
            target_prompt_parts.append(f"# Training readout\n{diagnostic_instruction.strip()}")
        target_user_prompt = "\n\n".join(target_prompt_parts)
        try:
            from skillopt.envs.spreadsheetbench.react_agent import _build_system
            target_system_prompt = _build_system(skill_content)
        except Exception:
            target_system_prompt = ""
        if target_system_prompt:
            with open(os.path.join(task_out_dir, "target_system_prompt.txt"), "w") as f:
                f.write(target_system_prompt)
            result["target_system_prompt"] = target_system_prompt
        with open(os.path.join(task_out_dir, "target_user_prompt.txt"), "w") as f:
            f.write(target_user_prompt)
        result["target_user_prompt"] = target_user_prompt

        # ── Stage 1: run ReAct agent on test case 1 ─────────────────────
        result["phase"] = "agent"

        work_dir = tempfile.mkdtemp(prefix=f"react_{task_id}_")
        try:
            # Copy input so agent works in an isolated directory
            work_input = os.path.join(work_dir, os.path.basename(ip1))
            shutil.copy2(ip1, work_input)

            agent_result = run_react(
                instruction=instruction,
                input_path=work_input,
                output_path=pred_path_1,
                work_dir=work_dir,
                instruction_type=instruction_type,
                answer_position=answer_position_eval,
                skill_content=skill_content,
                max_turns=max_turns,
                max_output_tokens=max_completion_tokens,
                diagnostic_mode=diagnostic_mode,
                diagnostic_instruction=diagnostic_instruction,
                diagnostic_trace_context=diagnostic_trace_context,
            )
            result["n_turns"] = agent_result.get("n_turns", 0)
            if agent_result.get("target_system_prompt"):
                with open(os.path.join(task_out_dir, "target_system_prompt.txt"), "w") as f:
                    f.write(agent_result["target_system_prompt"])
                result["target_system_prompt"] = agent_result["target_system_prompt"]
            if agent_result.get("target_user_prompt"):
                with open(os.path.join(task_out_dir, "target_user_prompt.txt"), "w") as f:
                    f.write(agent_result["target_user_prompt"])
                result["target_user_prompt"] = agent_result["target_user_prompt"]

            # Save conversation log
            with open(os.path.join(task_out_dir, "conversation.json"), "w") as f:
                json.dump(
                    agent_result.get("conversation", []),
                    f, ensure_ascii=False, indent=2,
                )

            # Copy solution.py if the agent wrote one
            solution_src = os.path.join(work_dir, "solution.py")
            solution_dst = os.path.join(task_out_dir, "solution.py")
            if os.path.exists(solution_src):
                shutil.copy2(solution_src, solution_dst)

        except Exception as e:
            result["fail_reason"] = f"agent-error: {type(e).__name__}: {e}"
            result["error"] = traceback.format_exc()
            return result
        finally:
            shutil.rmtree(work_dir, ignore_errors=True)

        result["agent_ok"] = True

        # ── Stage 2: evaluate all test cases ─────────────────────────────
        result["phase"] = "eval"
        solution_path = os.path.join(task_out_dir, "solution.py")
        all_exec = True

        for i, (no, ip, ap) in enumerate(cases):
            pred_path = os.path.join(task_out_dir, f"{no}_pred.xlsx")

            if i > 0:
                # Re-apply solution.py to subsequent test cases
                if not os.path.exists(solution_path):
                    all_exec = False
                    result["cases"].append(
                        {"no": no, "stage": "exec", "ok": False, "error": "no-solution-py"}
                    )
                    if not result["fail_reason"]:
                        result["fail_reason"] = "no-solution-py-for-other-cases"
                    continue

                with open(solution_path) as f:
                    code = f.read()

                # Prepend new INPUT_PATH / OUTPUT_PATH
                preamble = (
                    f"INPUT_PATH  = {ip!r}\n"
                    f"OUTPUT_PATH = {pred_path!r}\n"
                )
                full_code = preamble + code

                ok_exec, err = run_generated_code(full_code, ip, pred_path)
                if not ok_exec:
                    all_exec = False
                    result["cases"].append(
                        {"no": no, "stage": "exec", "ok": False, "error": err[:500]}
                    )
                    if not result["fail_reason"]:
                        tail = err.strip().splitlines()[-1][:200] if err.strip() else "unknown"
                        result["fail_reason"] = f"exec-error: {tail}"
                    continue

            # ── Evaluate ─────────────────────────────────────────────────
            if not os.path.exists(pred_path):
                all_exec = False
                result["cases"].append(
                    {"no": no, "stage": "exec", "ok": False, "error": "output-not-found"}
                )
                if not result["fail_reason"]:
                    result["fail_reason"] = "output-not-found"
                continue

            result["n_exec_pass"] += 1
            try:
                ev = evaluate(pred_path, ap, instruction_type, answer_position_eval)
            except Exception as e:  # noqa: BLE001
                ev = {"ok": False, "reason": f"eval-exception: {type(e).__name__}: {e}"}

            if ev["ok"]:
                result["n_pass"] += 1
            else:
                if not result["fail_reason"]:
                    result["fail_reason"] = f"eval-mismatch: {ev['reason'][:200]}"
            result["cases"].append(
                {"no": no, "stage": "eval", "ok": ev["ok"], "reason": ev.get("reason", "")}
            )

        result["exec_ok"] = all_exec
        n_cases = result["n_cases"]
        n_pass = result["n_pass"]
        result["soft"] = (n_pass / n_cases) if n_cases else 0.0
        result["hard"] = 1 if (n_cases > 0 and n_pass == n_cases) else 0
        result["ok"] = bool(result["hard"])
        if result["ok"]:
            result["fail_reason"] = ""
        return result

    except Exception as e:  # noqa: BLE001
        result["fail_reason"] = f"unexpected: {type(e).__name__}: {e}"
        result["error"] = traceback.format_exc()
        return result


# ── Batch runner ─────────────────────────────────────────────────────────────


def run_spreadsheet_batch(
    items: list[dict],
    data_root: str,
    out_root: str,
    skill_content: str,
    max_turns: int = 30,
    max_completion_tokens: int = 16384,
    max_api_workers: int = 64,
    task_timeout: int = 600,
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context_by_id: dict[str, str] | None = None,
) -> list[dict]:
    """Run the ReAct agent on all items with ThreadPoolExecutor.

    Returns list of result dicts compatible with ``compute_score()``.
    """
    os.makedirs(out_root, exist_ok=True)

    # Check for already-done items (resume support)
    results_path = os.path.join(out_root, "results.jsonl")
    done_ids: set[str] = set()
    existing: list[dict] = []
    if os.path.exists(results_path):
        with open(results_path) as f:
            for line in f:
                try:
                    r = json.loads(line)
                    done_ids.add(str(r["id"]))
                    existing.append(r)
                except Exception:
                    pass

    pending = [it for it in items if str(it["id"]) not in done_ids]
    print(
        f"  [spreadsheet rollout] total={len(items)} done={len(done_ids)} "
        f"pending={len(pending)} workers={max_api_workers} task_timeout={task_timeout}s"
    )

    if not pending:
        return existing

    t0 = time.time()
    results = list(existing)
    started_at: dict[str, float] = {}

    def _timeout_result(item: dict) -> dict:
        return {
            "id": str(item["id"]),
            "ok": False,
            "phase": "timeout",
            "fail_reason": f"task-timeout-{task_timeout}s",
            "n_cases": 0, "n_pass": 0, "soft": 0.0, "hard": 0,
            "n_turns": 0, "cases": [], "error": "timeout",
        }

    def _error_result(item: dict, exc: Exception) -> dict:
        return {
            "id": str(item["id"]),
            "ok": False,
            "phase": "error",
            "fail_reason": f"unexpected: {type(exc).__name__}: {exc}",
            "n_cases": 0, "n_pass": 0, "soft": 0.0, "hard": 0,
            "n_turns": 0, "cases": [], "error": str(exc),
        }

    def _run_one(it: dict) -> dict:
        started_at[str(it["id"])] = time.time()
        return process_one(
            it,
            data_root,
            out_root,
            skill_content,
            max_turns,
            diagnostic_mode,
            diagnostic_instruction,
            (diagnostic_trace_context_by_id or {}).get(str(it["id"]), ""),
            max_completion_tokens,
        )

    ex = ThreadPoolExecutor(max_workers=max_api_workers)
    try:
        futs = {ex.submit(_run_one, it): it for it in pending}
        pending_futs = set(futs)
        finished = 0
        while pending_futs:
            done, _ = wait(pending_futs, timeout=5, return_when=FIRST_COMPLETED)
            now = time.time()
            timed_out = [
                fut for fut in pending_futs - done
                if str(futs[fut]["id"]) in started_at
                and now - started_at[str(futs[fut]["id"])] >= task_timeout
            ]
            for fut in done:
                pending_futs.remove(fut)
                item = futs[fut]
                try:
                    res = fut.result()
                except FuturesTimeoutError:
                    res = _timeout_result(item)
                except Exception as e:  # noqa: BLE001
                    res = _error_result(item, e)
                results.append(res)
                finished += 1
                status = "PASS" if res.get("hard") else ("TIMEOUT" if res.get("phase") == "timeout" else "FAIL")
                dt = time.time() - t0
                print(
                    f"    {finished}/{len(pending)} id={res['id']:<10} {status}  "
                    f"turns={res.get('n_turns', 0):<3} "
                    f"cases={res.get('n_pass', 0)}/{res.get('n_cases', 0)}  "
                    f"dt={dt:.0f}s"
                )
            for fut in timed_out:
                pending_futs.remove(fut)
                res = _timeout_result(futs[fut])
                results.append(res)
                finished += 1
                status = "TIMEOUT"
                dt = time.time() - t0
                print(
                    f"    {finished}/{len(pending)} id={res['id']:<10} {status}  "
                    f"turns={res.get('n_turns', 0):<3} "
                    f"cases={res.get('n_pass', 0)}/{res.get('n_cases', 0)}  "
                    f"dt={dt:.0f}s"
                )
    finally:
        ex.shutdown(wait=False, cancel_futures=True)

    return results


# ── Codegen per-task worker (no tool-call) ──────────────────────────────────


def process_one_codegen(
    item: dict,
    data_root: str,
    out_root: str,
    skill_content: str,
    mode: str = "single",
    max_turns: int = 5,
    max_completion_tokens: int = 16384,
    task_timeout: int = 600,
    use_eval_feedback: bool = False,
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
) -> dict:
    """Run codegen agent (single or multi-round) on one SpreadsheetBench task.

    This matches the official evaluation setting: LLM generates a Python code
    block, no function-calling / tool-use.
    """
    from skillopt.envs.spreadsheetbench.codegen_agent import run_single, run_multi

    task_id = str(item["id"])
    instruction = item["instruction"]
    instruction_type = item.get("instruction_type", "")
    answer_position = item.get("answer_position", "")
    answer_sheet = item.get("answer_sheet", "")
    if answer_position and answer_sheet and "!" not in answer_position:
        answer_position_eval = f"{answer_sheet}!{answer_position}"
    else:
        answer_position_eval = answer_position

    itype_lower = (instruction_type or "").lower()
    if "cell" in itype_lower:
        task_type = "cell_level"
    elif "sheet" in itype_lower:
        task_type = "sheet_level"
    else:
        task_type = "other"

    sp = item.get("spreadsheet_path", f"spreadsheet/{task_id}")
    task_dir = sp if os.path.isabs(sp) else os.path.join(data_root, sp)

    result = {
        "id": task_id,
        "ok": False,
        "instruction_type": instruction_type,
        "task_type": task_type,
        "task_description": instruction,
        "phase": "setup",
        "fail_reason": "",
        "llm_ok": False,
        "code_ok": False,
        "exec_ok": False,
        "n_cases": 0,
        "n_exec_pass": 0,
        "n_pass": 0,
        "soft": 0.0,
        "hard": 0,
        "n_turns": 0,
        "cases": [],
        "error": "",
    }

    try:
        cases = _find_test_cases(task_dir)
        result["n_cases"] = len(cases)
        if not cases:
            result["fail_reason"] = "no-test-cases"
            return result

        task_out_dir = os.path.join(out_root, "predictions", task_id)
        os.makedirs(task_out_dir, exist_ok=True)

        # ── Save context for Optimizer (Reflect stage) ──────────────────
        from skillopt.envs.spreadsheetbench.codegen_agent import (
            _preview_workbook, _build_system, _build_user,
        )
        first_input_for_preview = cases[0][1]
        try:
            preview_text = _preview_workbook(first_input_for_preview)
        except Exception:
            preview_text = "(preview failed)"
        target_system = _build_system(skill_content)
        target_user = _build_user(
            instruction,
            first_input_for_preview,
            instruction_type,
            answer_position_eval,
            diagnostic_mode=diagnostic_mode,
            diagnostic_instruction=diagnostic_instruction,
            diagnostic_trace_context=diagnostic_trace_context,
        )

        with open(os.path.join(task_out_dir, "spreadsheet_preview.txt"), "w") as f:
            f.write(preview_text)
        with open(os.path.join(task_out_dir, "target_system_prompt.txt"), "w") as f:
            f.write(target_system)
        with open(os.path.join(task_out_dir, "target_user_prompt.txt"), "w") as f:
            f.write(target_user)

        result["spreadsheet_preview"] = preview_text
        result["target_system_prompt"] = target_system
        result["target_user_prompt"] = target_user

        # ── LLM phase ──────────────────────────────────────────────────
        result["phase"] = "llm"
        first_input = cases[0][1]
        first_gold = cases[0][2]
        first_pred = os.path.join(task_out_dir, f"{cases[0][0]}_pred.xlsx")

        try:
            if mode == "multi":
                agent_result = run_multi(
                    instruction=instruction,
                    input_xlsx=first_input,
                    output_path=first_pred,
                    instruction_type=instruction_type,
                    answer_position=answer_position_eval,
                    skill_content=skill_content,
                    max_turns=max_turns,
                    max_output_tokens=max_completion_tokens,
                    task_timeout=task_timeout,
                    gold_path=first_gold if use_eval_feedback else "",
                    diagnostic_mode=diagnostic_mode,
                    diagnostic_instruction=diagnostic_instruction,
                    diagnostic_trace_context=diagnostic_trace_context,
                )
            else:
                agent_result = run_single(
                    instruction=instruction,
                    input_xlsx=first_input,
                    output_path=first_pred,
                    instruction_type=instruction_type,
                    answer_position=answer_position_eval,
                    skill_content=skill_content,
                    max_output_tokens=max_completion_tokens,
                    task_timeout=task_timeout,
                    diagnostic_mode=diagnostic_mode,
                    diagnostic_instruction=diagnostic_instruction,
                    diagnostic_trace_context=diagnostic_trace_context,
                )
        except Exception as e:  # noqa: BLE001
            result["fail_reason"] = f"llm-call-failed: {type(e).__name__}: {e}"
            result["error"] = traceback.format_exc()
            return result

        result["llm_ok"] = True
        result["n_turns"] = agent_result.get("n_turns", 1)
        code = agent_result.get("code", "")
        raw = agent_result.get("raw", "")

        # Save artifacts
        with open(os.path.join(task_out_dir, "code.py"), "w") as f:
            f.write(code)
        with open(os.path.join(task_out_dir, "raw.txt"), "w") as f:
            f.write(raw)
        if agent_result.get("conversation"):
            with open(os.path.join(task_out_dir, "conversation.json"), "w") as f:
                json.dump(agent_result["conversation"], f, ensure_ascii=False, indent=2)

        if not code.strip():
            result["phase"] = "extract"
            result["fail_reason"] = "empty-code-block"
            return result
        result["code_ok"] = True

        # ── Exec + eval per test case ──────────────────────────────────
        result["phase"] = "exec"
        all_exec = True
        # Collect enrichment info for the conversation/trajectory
        enrichment_parts: list[str] = []

        for no, ip, ap in cases:
            pred_path = os.path.join(task_out_dir, f"{no}_pred.xlsx")

            # For multi mode, the first case may already be produced
            if not os.path.exists(pred_path):
                ok_exec, err = run_generated_code(code, ip, pred_path)
                if not ok_exec:
                    all_exec = False
                    result["cases"].append(
                        {"no": no, "stage": "exec", "ok": False, "error": err[:500]}
                    )
                    if not result["fail_reason"]:
                        tail = err.strip().splitlines()[-1][:200] if err.strip() else "unknown"
                        result["fail_reason"] = f"exec-error: {tail}"
                    enrichment_parts.append(
                        f"## Execution (case {no})\nERROR: {err[:500]}"
                    )
                    continue

            if not os.path.exists(pred_path):
                all_exec = False
                result["cases"].append(
                    {"no": no, "stage": "exec", "ok": False, "error": "output-not-found"}
                )
                if not result["fail_reason"]:
                    result["fail_reason"] = "output-not-found"
                continue

            result["n_exec_pass"] += 1
            try:
                ev = evaluate(pred_path, ap, instruction_type, answer_position_eval)
            except Exception as e:  # noqa: BLE001
                ev = {"ok": False, "reason": f"eval-exception: {type(e).__name__}: {e}"}

            if ev["ok"]:
                result["n_pass"] += 1
            else:
                if not result["fail_reason"]:
                    result["fail_reason"] = f"eval-mismatch: {ev['reason'][:200]}"
            result["cases"].append(
                {"no": no, "stage": "eval", "ok": ev["ok"], "reason": ev.get("reason", "")}
            )

            # Auto-verify: reopen output and compare cells at answer_position
            if answer_position_eval:
                verify_report = _auto_verify_output(pred_path, ap, answer_position_eval)
                enrichment_parts.append(
                    f"## Eval Result (case {no}): {'PASS' if ev['ok'] else 'FAIL'}\n"
                    f"{ev.get('reason', '')}\n\n{verify_report}"
                )

        result["exec_ok"] = all_exec

        # ── Enrich conversation with eval details ──────────────────────
        if enrichment_parts:
            enrichment_msg = "\n\n---\n\n".join(enrichment_parts)
            conversation = agent_result.get("conversation", [])
            conversation.append({
                "role": "system",
                "content": f"[POST-EXECUTION VERIFICATION]\n\n{enrichment_msg}",
            })
            # Re-save the enriched conversation
            with open(os.path.join(task_out_dir, "conversation.json"), "w") as f:
                json.dump(conversation, f, ensure_ascii=False, indent=2)
        n_cases = result["n_cases"]
        n_pass = result["n_pass"]
        result["soft"] = (n_pass / n_cases) if n_cases else 0.0
        result["hard"] = 1 if (n_cases > 0 and n_pass == n_cases) else 0
        result["ok"] = bool(result["hard"])
        if result["ok"]:
            result["fail_reason"] = ""
        return result

    except Exception as e:  # noqa: BLE001
        result["fail_reason"] = f"unexpected: {type(e).__name__}: {e}"
        result["error"] = traceback.format_exc()
        return result


# ── Codegen batch runner ────────────────────────────────────────────────────


def run_spreadsheet_batch_codegen(
    items: list[dict],
    data_root: str,
    out_root: str,
    skill_content: str,
    mode: str = "single",
    max_turns: int = 5,
    max_completion_tokens: int = 16384,
    max_api_workers: int = 32,
    task_timeout: int = 0,
    use_eval_feedback: bool = False,
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context_by_id: dict[str, str] | None = None,
) -> list[dict]:
    """Run codegen agent on all items (no tool-call).

    Args:
        mode: "single" or "multi".
        task_timeout: Hard per-task timeout in seconds at the future level.
            0 or negative disables the per-task timeout.
    """
    no_task_timeout = task_timeout <= 0
    task_timeout_label = "none" if no_task_timeout else f"{task_timeout}s"

    os.makedirs(out_root, exist_ok=True)

    results_path = os.path.join(out_root, "results.jsonl")
    done_ids: set[str] = set()
    existing: list[dict] = []
    if os.path.exists(results_path):
        with open(results_path) as f:
            for line in f:
                try:
                    r = json.loads(line)
                    done_ids.add(str(r["id"]))
                    existing.append(r)
                except Exception:
                    pass

    pending = [it for it in items if str(it["id"]) not in done_ids]
    print(
        f"  [spreadsheet codegen-{mode}] total={len(items)} done={len(done_ids)} "
        f"pending={len(pending)} workers={max_api_workers} task_timeout={task_timeout_label}"
    )

    if not pending:
        return existing

    t0 = time.time()
    results = list(existing)

    started_at: dict[str, float] = {}

    def _run_one(it: dict) -> dict:
        started_at[str(it["id"])] = time.time()
        return process_one_codegen(
            it,
            data_root,
            out_root,
            skill_content,
            mode,
            max_turns,
            max_completion_tokens,
            task_timeout,
            use_eval_feedback,
            diagnostic_mode,
            diagnostic_instruction,
            (diagnostic_trace_context_by_id or {}).get(str(it["id"]), ""),
        )

    def _timeout_result(item: dict) -> dict:
        return {
            "id": str(item["id"]),
            "ok": False,
            "instruction_type": item.get("instruction_type", ""),
            "task_type": "other",
            "phase": "timeout",
            "fail_reason": f"task-timeout-{task_timeout}s",
            "n_cases": 0, "n_pass": 0, "soft": 0.0, "hard": 0,
            "n_turns": 0, "cases": [], "error": "timeout",
        }

    def _error_result(item: dict, e: Exception) -> dict:
        return {
            "id": str(item["id"]),
            "ok": False,
            "instruction_type": item.get("instruction_type", ""),
            "task_type": "other",
            "phase": "error",
            "fail_reason": f"unexpected: {type(e).__name__}: {e}",
            "n_cases": 0, "n_pass": 0, "soft": 0.0, "hard": 0,
            "n_turns": 0, "cases": [], "error": str(e),
        }

    def _record(res: dict, i: int) -> None:
        results.append(res)
        status = "PASS" if res.get("hard") else ("TIMEOUT" if res.get("phase") == "timeout" else "FAIL")
        dt = time.time() - t0
        print(
            f"    {i}/{len(pending)} id={res['id']:<10} {status}  "
            f"turns={res.get('n_turns', 0):<3} "
            f"cases={res.get('n_pass', 0)}/{res.get('n_cases', 0)}  "
            f"dt={dt:.0f}s"
        )

    ex = ThreadPoolExecutor(max_workers=max_api_workers)
    try:
        futs = {ex.submit(_run_one, it): it for it in pending}
        pending_futs = set(futs)
        finished = 0
        while pending_futs:
            done, _ = wait(pending_futs, timeout=5, return_when=FIRST_COMPLETED)
            now = time.time()
            timed_out = [] if no_task_timeout else [
                fut for fut in pending_futs - done
                if str(futs[fut]["id"]) in started_at
                and now - started_at[str(futs[fut]["id"])] >= task_timeout
            ]
            for fut in done:
                pending_futs.remove(fut)
                item = futs[fut]
                try:
                    res = fut.result()
                except FuturesTimeoutError:
                    res = _timeout_result(item)
                except Exception as e:  # noqa: BLE001
                    res = _error_result(item, e)
                finished += 1
                _record(res, finished)
            for fut in timed_out:
                pending_futs.remove(fut)
                fut.cancel()
                finished += 1
                _record(_timeout_result(futs[fut]), finished)
    finally:
        ex.shutdown(wait=False, cancel_futures=True)

    return results
